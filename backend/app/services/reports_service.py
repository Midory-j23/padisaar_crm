import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.account import Account, Industry, OrgSize, PriorityLevel, RelationshipStatus
from app.models.activity import Activity, ActivityType
from app.models.audit_log import AuditAction, AuditLog
from app.models.contact import Contact, InfluenceLevel, Sentiment
from app.models.opportunity import Opportunity, SalesStage
from app.models.user import User
from app.schemas.reports import (
    AuditLogEntryResponse,
    ImportConfirmResponse,
    ImportPreviewResponse,
    ImportRowPreview,
)
from app.services import account_service, activity_service, opportunity_service
from app.services.exceptions import BadRequestError
from app.utils.audit import entity_summary, log_audit
from app.utils.datetime_utils import to_naive_utc
from app.utils.jalali import to_jalali_str
from app.utils.pdf_export import build_table_pdf

HEADER_FILL = PatternFill("solid", fgColor="1e3a5f")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri")

INDUSTRY_LABELS = {
    Industry.OIL_GAS: "نفت و گاز",
    Industry.PETROCHEMICAL: "پتروشیمی",
    Industry.STEEL: "فولاد",
    Industry.MINING: "معدن",
    Industry.INFRASTRUCTURE: "زیرساخت",
    Industry.OTHER: "سایر",
}
SIZE_LABELS = {
    OrgSize.SMALL: "کوچک",
    OrgSize.MEDIUM: "متوسط",
    OrgSize.LARGE: "بزرگ",
}
PRIORITY_LABELS = {
    PriorityLevel.A_STRATEGIC: "A - استراتژیک",
    PriorityLevel.B_MEDIUM: "B - متوسط",
    PriorityLevel.C_GENERAL: "C - عمومی",
}
RELATIONSHIP_LABELS = {
    RelationshipStatus.CURRENT_CLIENT: "مشتری فعلی",
    RelationshipStatus.FORMER_CLIENT: "مشتری سابق",
    RelationshipStatus.NEW_LEAD: "سرنخ جدید",
    RelationshipStatus.COMPETITOR: "رقیب",
}
INFLUENCE_LABELS = {
    InfluenceLevel.DECISION_MAKER: "تصمیم‌گیر نهایی",
    InfluenceLevel.TECHNICAL_INFLUENCER: "تأثیرگذار فنی",
    InfluenceLevel.BLOCKER: "بازدارنده",
    InfluenceLevel.BUYER: "خریدار",
}
SENTIMENT_LABELS = {
    Sentiment.CHAMPION: "حامی",
    Sentiment.NEUTRAL: "خنثی",
    Sentiment.OPPONENT: "مخالف",
}
STAGE_LABELS = {
    SalesStage.INITIAL_CONTACT: "ارتباط اولیه",
    SalesStage.NEEDS_ASSESSMENT: "ارزیابی نیاز",
    SalesStage.PROPOSAL_SENT: "ارسال پروپوزال",
    SalesStage.NEGOTIATION: "مذاکره",
    SalesStage.CONTRACT_SIGNED: "عقد قرارداد",
    SalesStage.CLOSED_WON: "بسته شده - موفق",
    SalesStage.CLOSED_LOST: "بسته شده - ناموفق",
    SalesStage.ABANDONED: "متوقف شده",
}
ACTIVITY_TYPE_LABELS = {
    ActivityType.IN_PERSON_MEETING: "جلسه حضوری",
    ActivityType.PHONE_CALL: "تماس تلفنی",
    ActivityType.SITE_VISIT: "بازدید سایت",
    ActivityType.PROPOSAL_SENT: "ارسال پروپوزال",
    ActivityType.EMAIL: "ایمیل",
}
AUDIT_ACTION_LABELS = {
    AuditAction.CREATE: "ایجاد",
    AuditAction.UPDATE: "ویرایش",
    AuditAction.DELETE: "حذف",
}

REVERSE_INDUSTRY = {v: k for k, v in INDUSTRY_LABELS.items()}
REVERSE_INDUSTRY.update({k.value: k for k in Industry})
REVERSE_SIZE = {v: k for k, v in SIZE_LABELS.items()}
REVERSE_SIZE.update({k.value: k for k in OrgSize})
REVERSE_PRIORITY = {v: k for k, v in PRIORITY_LABELS.items()}
REVERSE_PRIORITY.update({k.value: k for k in PriorityLevel})
REVERSE_RELATIONSHIP = {v: k for k, v in RELATIONSHIP_LABELS.items()}
REVERSE_RELATIONSHIP.update({k.value: k for k in RelationshipStatus})
REVERSE_INFLUENCE = {v: k for k, v in INFLUENCE_LABELS.items()}
REVERSE_INFLUENCE.update({k.value: k for k in InfluenceLevel})
REVERSE_SENTIMENT = {v: k for k, v in SENTIMENT_LABELS.items()}
REVERSE_SENTIMENT.update({k.value: k for k in Sentiment})


def _label(mapping: dict, value) -> str:
    if value is None:
        return ""
    if value in mapping:
        return mapping[value]
    return str(value)


def _build_workbook(title: str, headers: list[str], rows: list[list[Any]]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.sheet_view.rightToLeft = True

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="right")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


async def export_accounts(db: AsyncSession, template: bool = False) -> io.BytesIO:
    headers = [
        "نام سازمان",
        "شناسه ملی",
        "صنعت",
        "اندازه",
        "سطح اولویت",
        "وضعیت ارتباط",
        "موقعیت",
        "وبسایت",
        "کارشناس پیگیر",
        "تاریخ ثبت",
    ]
    rows: list[list[Any]] = []
    if not template:
        result = await db.execute(
            select(Account).options(selectinload(Account.account_manager)).order_by(Account.name)
        )
        for acc in result.scalars().all():
            rows.append(
                [
                    acc.name,
                    acc.national_id or "",
                    _label(INDUSTRY_LABELS, acc.industry),
                    _label(SIZE_LABELS, acc.size),
                    _label(PRIORITY_LABELS, acc.priority_level),
                    _label(RELATIONSHIP_LABELS, acc.relationship_status),
                    acc.location or "",
                    acc.website or "",
                    acc.account_manager.name if acc.account_manager else "",
                    to_jalali_str(acc.created_at),
                ]
            )
    return _build_workbook("سازمان‌ها", headers, rows)


async def export_contacts(db: AsyncSession, template: bool = False) -> io.BytesIO:
    headers = [
        "نام و نام خانوادگی",
        "سازمان",
        "سمت",
        "دپارتمان",
        "موبایل",
        "خط مستقیم",
        "ایمیل",
        "سطح اثرگذاری",
        "تمایل",
        "تاریخ ثبت",
    ]
    rows: list[list[Any]] = []
    if not template:
        result = await db.execute(
            select(Contact).options(selectinload(Contact.account)).order_by(Contact.full_name)
        )
        for c in result.scalars().all():
            rows.append(
                [
                    c.full_name,
                    c.account.name if c.account else "",
                    c.job_title or "",
                    c.department or "",
                    c.mobile,
                    c.direct_line or "",
                    c.email or "",
                    _label(INFLUENCE_LABELS, c.influence_level),
                    _label(SENTIMENT_LABELS, c.sentiment),
                    to_jalali_str(c.created_at),
                ]
            )
    return _build_workbook("مخاطبان", headers, rows)


async def export_opportunities(db: AsyncSession, template: bool = False) -> io.BytesIO:
    headers = [
        "عنوان",
        "سازمان",
        "نوع پروژه",
        "مرحله فروش",
        "ارزش تخمینی",
        "احتمال",
        "منبع سرنخ",
        "تاریخ بسته‌شدن",
        "کارشناس",
        "تاریخ ثبت",
    ]
    rows: list[list[Any]] = []
    if not template:
        result = await db.execute(
            select(Opportunity)
            .options(selectinload(Opportunity.account), selectinload(Opportunity.assigned_to))
            .order_by(Opportunity.updated_at.desc())
        )
        for o in result.scalars().all():
            rows.append(
                [
                    o.title,
                    o.account.name if o.account else "",
                    o.project_type.value if o.project_type else "",
                    _label(STAGE_LABELS, o.sales_stage),
                    float(o.estimated_value) if o.estimated_value else "",
                    o.probability,
                    o.lead_source.value if o.lead_source else "",
                    to_jalali_str(o.expected_close_date),
                    o.assigned_to.name if o.assigned_to else "",
                    to_jalali_str(o.created_at),
                ]
            )
    return _build_workbook("فرصت‌ها", headers, rows)


async def export_activities(db: AsyncSession, template: bool = False) -> io.BytesIO:
    headers = [
        "نوع فعالیت",
        "تاریخ",
        "سازمان",
        "فرصت",
        "مخاطب",
        "نتیجه",
        "اقدام بعدی",
        "تاریخ پیگیری",
        "ثبت‌کننده",
        "تاریخ ثبت",
    ]
    rows: list[list[Any]] = []
    if not template:
        result = await db.execute(
            select(Activity)
            .options(
                selectinload(Activity.account),
                selectinload(Activity.opportunity),
                selectinload(Activity.contact),
                selectinload(Activity.contacts),
                selectinload(Activity.created_by),
            )
            .order_by(Activity.activity_date.desc())
        )
        for a in result.scalars().unique().all():
            contact_label = "، ".join(c.full_name for c in a.contacts) if a.contacts else (
                a.contact.full_name if a.contact else ""
            )
            rows.append(
                [
                    _label(ACTIVITY_TYPE_LABELS, a.activity_type),
                    to_jalali_str(a.activity_date),
                    a.account.name if a.account else "",
                    a.opportunity.title if a.opportunity else "",
                    contact_label,
                    a.outcome or "",
                    a.next_step or "",
                    to_jalali_str(a.follow_up_date),
                    a.created_by.name if a.created_by else "",
                    to_jalali_str(a.created_at),
                ]
            )
    return _build_workbook("فعالیت‌ها", headers, rows)


async def report_opportunities(
    db: AsyncSession,
    user: User,
    *,
    stage: SalesStage | None = None,
    account_id: str | None = None,
    assigned_to_id: str | None = None,
    search: str | None = None,
) -> dict:
    data = await opportunity_service.list_opportunities(
        db,
        user,
        stage=stage,
        account_id=account_id,
        assigned_to_id=assigned_to_id,
        search=search,
        page=1,
        per_page=10000,
    )
    return {"items": data["items"], "total": data["total"]}


async def report_activities(
    db: AsyncSession,
    user: User,
    *,
    account_id: str | None = None,
    opportunity_id: str | None = None,
    activity_type: ActivityType | None = None,
    assigned_to: str | None = None,
    from_date: datetime | None = None,
    to_date: datetime | None = None,
) -> dict:
    data = await activity_service.list_activities(
        db,
        user,
        account_id=account_id,
        opportunity_id=opportunity_id,
        activity_type=activity_type,
        assigned_to=assigned_to,
        from_date=from_date,
        to_date=to_date,
        page=1,
        per_page=10000,
    )
    return {"items": data["items"], "total": data["total"]}


def _parse_enum(value: str | None, reverse_map: dict) -> Any:
    if not value:
        return None
    key = str(value).strip()
    return reverse_map.get(key)


def _cell_str(value) -> str | None:
    if value is None:
        return None
    return str(value).strip() or None


async def preview_import_accounts(db: AsyncSession, content: bytes) -> ImportPreviewResponse:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise BadRequestError("فایل Excel معتبر نیست") from exc

    ws = wb.active
    rows_preview: list[ImportRowPreview] = []
    seen_national_ids: set[str] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        name = _cell_str(row[0] if len(row) > 0 else None)
        national_id_raw = row[1] if len(row) > 1 else None
        national_id = None
        if national_id_raw is not None:
            national_id = str(int(national_id_raw)) if isinstance(national_id_raw, (int, float)) else str(national_id_raw).strip()

        record = {
            "name": name,
            "national_id": national_id,
            "industry": _parse_enum(_cell_str(row[2] if len(row) > 2 else None), REVERSE_INDUSTRY),
            "size": _parse_enum(_cell_str(row[3] if len(row) > 3 else None), REVERSE_SIZE),
            "priority_level": _parse_enum(_cell_str(row[4] if len(row) > 4 else None), REVERSE_PRIORITY),
            "relationship_status": _parse_enum(_cell_str(row[5] if len(row) > 5 else None), REVERSE_RELATIONSHIP),
            "location": _cell_str(row[6] if len(row) > 6 else None),
            "website": _cell_str(row[7] if len(row) > 7 else None),
        }

        errors: list[str] = []
        if not name:
            errors.append("نام سازمان الزامی است")
        if national_id and (not national_id.isdigit() or len(national_id) != 11):
            errors.append("شناسه ملی باید ۱۱ رقم باشد")
        if national_id:
            if national_id in seen_national_ids:
                errors.append("شناسه ملی تکراری در فایل")
            seen_national_ids.add(national_id)
            dup = await db.execute(select(Account).where(Account.national_id == national_id))
            if dup.scalar_one_or_none():
                errors.append("شناسه ملی قبلاً در سیستم ثبت شده")

        enum_fields = [
            ("industry", record["industry"], row[2] if len(row) > 2 else None),
            ("size", record["size"], row[3] if len(row) > 3 else None),
            ("priority_level", record["priority_level"], row[4] if len(row) > 4 else None),
            ("relationship_status", record["relationship_status"], row[5] if len(row) > 5 else None),
        ]
        for field_name, parsed, raw in enum_fields:
            if raw and str(raw).strip() and parsed is None:
                errors.append(f"مقدار نامعتبر برای {field_name}")

        serializable = {
            **record,
            "industry": record["industry"].value if record["industry"] else None,
            "size": record["size"].value if record["size"] else None,
            "priority_level": record["priority_level"].value if record["priority_level"] else None,
            "relationship_status": record["relationship_status"].value if record["relationship_status"] else None,
        }

        rows_preview.append(
            ImportRowPreview(
                row_number=row_idx,
                record=serializable,
                errors=errors,
                valid=len(errors) == 0,
            )
        )

    wb.close()
    valid_count = sum(1 for r in rows_preview if r.valid)
    return ImportPreviewResponse(
        rows=rows_preview,
        valid_count=valid_count,
        error_count=len(rows_preview) - valid_count,
    )


async def confirm_import_accounts(
    db: AsyncSession, user: User, records: list[dict[str, Any]]
) -> ImportConfirmResponse:
    if not records:
        raise BadRequestError("رکوردی برای ورود وجود ندارد")

    def _enum(enum_cls, value):
        if value is None:
            return None
        if isinstance(value, enum_cls):
            return value
        return enum_cls(value)

    created = 0
    for record in records:
        account = Account(
            name=record["name"],
            national_id=record.get("national_id"),
            industry=_enum(Industry, record.get("industry")),
            size=_enum(OrgSize, record.get("size")),
            priority_level=_enum(PriorityLevel, record.get("priority_level")),
            relationship_status=_enum(RelationshipStatus, record.get("relationship_status")),
            location=record.get("location"),
            website=record.get("website"),
        )
        db.add(account)
        await db.flush()
        await log_audit(db, "Account", account.id, AuditAction.CREATE, user.id, record)
        created += 1

    await db.commit()
    return ImportConfirmResponse(
        created_count=created,
        message=f"{created} سازمان با موفقیت وارد شد",
    )


async def preview_import_contacts(db: AsyncSession, content: bytes) -> ImportPreviewResponse:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise BadRequestError("فایل Excel معتبر نیست") from exc

    ws = wb.active
    rows_preview: list[ImportRowPreview] = []
    seen_mobiles: set[str] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        account_name = _cell_str(row[1] if len(row) > 1 else None)
        mobile_raw = row[4] if len(row) > 4 else None
        mobile = None
        if mobile_raw is not None:
            mobile = str(int(mobile_raw)) if isinstance(mobile_raw, (int, float)) else str(mobile_raw).strip()

        record = {
            "full_name": _cell_str(row[0] if len(row) > 0 else None),
            "account_name": account_name,
            "job_title": _cell_str(row[2] if len(row) > 2 else None),
            "department": _cell_str(row[3] if len(row) > 3 else None),
            "mobile": mobile,
            "direct_line": _cell_str(row[5] if len(row) > 5 else None),
            "email": _cell_str(row[6] if len(row) > 6 else None),
            "influence_level": _parse_enum(_cell_str(row[7] if len(row) > 7 else None), REVERSE_INFLUENCE),
            "sentiment": _parse_enum(_cell_str(row[8] if len(row) > 8 else None), REVERSE_SENTIMENT),
            "account_id": None,
        }

        errors: list[str] = []
        if not record["full_name"]:
            errors.append("نام مخاطب الزامی است")
        if not account_name:
            errors.append("نام سازمان الزامی است")
        if not mobile:
            errors.append("شماره موبایل الزامی است")
        elif mobile in seen_mobiles:
            errors.append("موبایل تکراری در فایل")
        if mobile:
            seen_mobiles.add(mobile)
            dup = await db.execute(select(Contact).where(Contact.mobile == mobile))
            if dup.scalar_one_or_none():
                errors.append("این موبایل قبلاً در سیستم ثبت شده")

        account = None
        if account_name:
            acc_result = await db.execute(select(Account).where(Account.name == account_name))
            account = acc_result.scalar_one_or_none()
            if not account:
                errors.append("سازمان یافت نشد — ابتدا سازمان را ثبت یا وارد کنید")
            else:
                record["account_id"] = account.id

        for field_name, parsed, raw in [
            ("influence_level", record["influence_level"], row[7] if len(row) > 7 else None),
            ("sentiment", record["sentiment"], row[8] if len(row) > 8 else None),
        ]:
            if raw and str(raw).strip() and parsed is None:
                errors.append(f"مقدار نامعتبر برای {field_name}")

        serializable = {
            **record,
            "influence_level": record["influence_level"].value if record["influence_level"] else None,
            "sentiment": record["sentiment"].value if record["sentiment"] else None,
        }

        rows_preview.append(
            ImportRowPreview(
                row_number=row_idx,
                record=serializable,
                errors=errors,
                valid=len(errors) == 0,
            )
        )

    wb.close()
    valid_count = sum(1 for r in rows_preview if r.valid)
    return ImportPreviewResponse(
        rows=rows_preview,
        valid_count=valid_count,
        error_count=len(rows_preview) - valid_count,
    )


async def confirm_import_contacts(
    db: AsyncSession, user: User, records: list[dict[str, Any]]
) -> ImportConfirmResponse:
    if not records:
        raise BadRequestError("رکوردی برای ورود وجود ندارد")

    def _enum(enum_cls, value):
        if value is None:
            return None
        if isinstance(value, enum_cls):
            return value
        return enum_cls(value)

    created = 0
    for record in records:
        contact = Contact(
            account_id=record["account_id"],
            full_name=record["full_name"],
            job_title=record.get("job_title"),
            department=record.get("department"),
            mobile=record["mobile"],
            direct_line=record.get("direct_line"),
            email=record.get("email"),
            influence_level=_enum(InfluenceLevel, record.get("influence_level")),
            sentiment=_enum(Sentiment, record.get("sentiment")),
        )
        db.add(contact)
        await db.flush()
        await log_audit(db, "Contact", contact.id, AuditAction.CREATE, user.id, record)
        created += 1

    await db.commit()
    return ImportConfirmResponse(
        created_count=created,
        message=f"{created} مخاطب با موفقیت وارد شد",
    )


async def export_opportunities_pdf(
    db: AsyncSession,
    user: User,
    *,
    stage: SalesStage | None = None,
    account_id: str | None = None,
    assigned_to_id: str | None = None,
    search: str | None = None,
) -> io.BytesIO:
    data = await report_opportunities(
        db, user, stage=stage, account_id=account_id, assigned_to_id=assigned_to_id, search=search
    )
    headers = ["عنوان", "سازمان", "مرحله", "ارزش", "احتمال", "کارشناس"]
    rows = [
        [
            o.title,
            o.account_name or "",
            _label(STAGE_LABELS, o.sales_stage),
            float(o.estimated_value) if o.estimated_value else "",
            o.probability,
            o.assigned_to_name or "",
        ]
        for o in data["items"]
    ]
    return build_table_pdf("گزارش فرصت‌ها", headers, rows)


async def export_activities_pdf(
    db: AsyncSession,
    user: User,
    *,
    account_id: str | None = None,
    opportunity_id: str | None = None,
    activity_type: ActivityType | None = None,
    assigned_to: str | None = None,
    from_date: datetime | None = None,
    to_date: datetime | None = None,
) -> io.BytesIO:
    data = await report_activities(
        db,
        user,
        account_id=account_id,
        opportunity_id=opportunity_id,
        activity_type=activity_type,
        assigned_to=assigned_to,
        from_date=from_date,
        to_date=to_date,
    )
    headers = ["نوع", "تاریخ", "سازمان", "فرصت", "مخاطب", "ثبت‌کننده"]
    rows = [
        [
            _label(ACTIVITY_TYPE_LABELS, a.activity_type),
            to_jalali_str(a.activity_date),
            a.account_name or "",
            a.opportunity_title or "",
            "، ".join(a.contact_names) if a.contact_names else (a.contact_name or ""),
            a.created_by_name or "",
        ]
        for a in data["items"]
    ]
    return build_table_pdf("گزارش فعالیت‌ها", headers, rows)


async def list_audit_logs(
    db: AsyncSession,
    *,
    entity_type: str | None = None,
    user_id: str | None = None,
    action: AuditAction | None = None,
    from_date: datetime | None = None,
    to_date: datetime | None = None,
    page: int = 1,
    per_page: int = 20,
) -> dict:
    from_date = to_naive_utc(from_date)
    to_date = to_naive_utc(to_date)
    query = select(AuditLog).options(selectinload(AuditLog.changed_by))

    if entity_type:
        query = query.where(AuditLog.entity_type == entity_type)
    if user_id:
        query = query.where(AuditLog.changed_by_id == user_id)
    if action:
        query = query.where(AuditLog.action == action)
    if from_date:
        query = query.where(AuditLog.created_at >= from_date)
    if to_date:
        query = query.where(AuditLog.created_at <= to_date)

    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar() or 0

    query = (
        query.order_by(AuditLog.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    result = await db.execute(query)
    logs = result.scalars().all()

    items = [
        AuditLogEntryResponse(
            id=log.id,
            entity_type=log.entity_type,
            entity_id=log.entity_id,
            entity_summary=entity_summary(log.entity_type, log.change_data or {}),
            action=log.action,
            changed_by_id=log.changed_by_id,
            changed_by_name=log.changed_by.name if log.changed_by else None,
            change_data=log.change_data or {},
            created_at=log.created_at,
        )
        for log in logs
    ]
    return {"items": items, "total": total, "page": page, "per_page": per_page}
